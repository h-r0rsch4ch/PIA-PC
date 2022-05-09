from openpyxl import Workbook
from virus_total_apis import PublicApi
import time
import socket
import logging


def analizarUrl(txt, apikey):
    logging.info('Análisis de urls iniciado')
    # Abrimos nuestro txt y un excel en blanco
    fo = open(txt, "r")
    libro = Workbook()
    hoja = libro.active
    # Declaramos el uso para cada columna
    hoja.cell(1, 1, "Url")
    hoja.cell(1, 2, "Fecha de análisis")
    hoja.cell(1, 3, "Total de análisis")
    hoja.cell(1, 4, "Análisis positivos")
    hoja.cell(1, 5, "Clasificación")
    # Obtenemos la api para VirusTotal
    api_key = apikey
    api = PublicApi(api_key)
    # Definimos nuestro protocolo de socket
    protocolo = socket.IPPROTO_TCP
    i = 2
    # Abrimos un txt para guardar los resultados de socket
    for line in fo:
        # Verificamos la conexión a cada url
        response = api.get_url_report(line)
        # Registramos la url
        hoja.cell(i, 1, line)
        if response["response_code"] == 200:
            # Registramos cada apartado en su columna
            hoja.cell(i, 2, response["results"]["scan_date"])
            hoja.cell(i, 3, len(response["results"]["scans"]))
            hoja.cell(i, 4, response["results"]["positives"])
            if response["results"]["positives"] <= 3:
                hoja.cell(i, 5, "Baja")
            elif (response["results"]["positives"] > 3 and
                response["results"]["positives"] <= 10):
                hoja.cell(i, 5, "Medio")
            elif response["results"]["positives"] > 10:
                hoja.cell(i, 5, "Alto")
        else:
            hoja.cell(i, 2, "Error de conexión")
        # Esperamos para poder seguir usando la api
        i += 1
        time.sleep(10)
    logging.info('Análisis de urls terminado')
    # Guardamos nuestros documentos
    fo.close()
    # foo.close()
    libro.save("reporte_analisis.xlsx")